#!/usr/bin/env python
# -*- coding: utf-8 -*-
from math import sqrt
import random
import read_file
import split
import timeitem as t
import build_graph as g

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

net_weights = g.Graph()


def scores_comp(org, list2, group_topic, members):  # 得分即为每个人对于活动的每个兴趣点的平均值
    scores = {}
    for id_member in list2:
        same = 0
        num = len(group_topic)
        if id_member in org:
            same = num
        for topic_member in group_topic:
            if topic_member in members[id_member]:
                same = same + members[id_member][topic_member]
        scores[id_member] = same / num
    return scores


def sims_comp(list1, list2, members, group_topics, pattern):
    sims = {}
    add = []
    for id_member in list(set(list1 + list2)):
        if id_member not in members:
            for topic in group_topics:
                members[id_member][topic] = 1
            add.append(id_member)
    avg = {}
    for id_member in list(set(list1 + list2)):
        num = 0
        avg[id_member] = 0
        for topic in group_topics:
            if topic in members[id_member]:
                avg[id_member] += members[id_member][topic]
                num += 1
        '''
        for topic in members[id_member]:
            avg[id_member] += members[id_member][topic]
            num += 1
        '''
        avg[id_member] /= (num if num != 0 else 1)
    for id_member2 in list2:
        sim = {}
        for id_member1 in list1:
            mul = 0
            len1 = 0
            len2 = 0
            for topic in group_topics:
                # for topic in members[id_member2]:
                if topic not in members[id_member1]:
                    tmp1 = avg[id_member1]
                else:
                    tmp1 = members[id_member1][topic]
                if topic not in members[id_member2]:
                    tmp2 = avg[id_member2]
                else:
                    tmp2 = members[id_member2][topic]
                mul += (tmp1 - avg[id_member1]) * (tmp2 - avg[id_member2])
                len1 += (tmp1 - avg[id_member1]) * (tmp1 - avg[id_member1])
                len2 += (tmp2 - avg[id_member2]) * (tmp2 - avg[id_member2])
            tmp = sqrt(len1 * len2)
            sim_tmp = mul / (tmp if tmp != 0 else 1)
            if pattern == 5:
                if sim_tmp > 0.5:
                    sim[id_member1] = 1
                else:
                    sim[id_member1] = 0
            else:
                sim[id_member1] = sim_tmp
        sims[id_member2] = sim
    for id_member in add:
        del members[id_member]
    # return sorted(sim.items(), key=lambda x: x[1], reverse=False)
    return sims


def predict(id_group, id_event, members, groups, events_val, pattern, member_events, events_item):
    yes = []
    maybe = []
    no = []
    list1 = events_val[id_group][id_event]['org']
    list2 = events_val[id_group][id_event]['yes'] + events_val[id_group][id_event]['maybe'] + \
            events_val[id_group][id_event]['no']
    scores_base = scores_comp(list1, list2, groups[id_group]['topic'], members)
    if pattern <= 1:  # 0 and 1
        scores = scores_base
    elif pattern == 2 or pattern == 4:  # 2 and 4
        members = t.members_update(members, member_events, events_item, list2, events_val[id_group][id_event]['time'])
        scores = scores_comp(list1, list2, groups[id_group]['topic'], members)
    sims = sims_comp(list2, list2, members, groups[id_group]['topic'], pattern)
    if pattern >= 6:  # 6 and 7
        scores = {}
        for id_member1 in list2:
            scores[id_member1] = 0
            sum = 0
            for id_member2 in list2:
                if id_member2 != id_member1:
                    if id_member1 != 'null' and id_member2 != 'null' and int(id_member1[1:]) in net_weights.vertList \
                            and int(id_member2[1:]) in net_weights.vertList \
                            and net_weights.vertList[int(id_member2[1:])] \
                            in net_weights.vertList[int(id_member1[1:])].connectedTo.keys():
                        scores[id_member1] += sims[id_member1][id_member2] * scores_base[id_member2]
                        # sum += scores_base[id_member2]
                        sum += sims[id_member1][id_member2]
            scores[id_member1] = scores_base[id_member1] + scores[id_member1] / (sum if sum != 0 else 1)
    elif pattern > 2:  # 3  4 and 5

        scores = {}
        for id_member1 in list2:
            scores[id_member1] = 0
            sum = 0
            for id_member2 in list2:
                if id_member2 != id_member1:
                    scores[id_member1] += sims[id_member1][id_member2] * scores_base[id_member2]
                    # sum += scores_base[id_member2]
                    sum += sims[id_member1][id_member2]
            scores[id_member1] = scores_base[id_member1] + scores[id_member1] / (sum if sum != 0 else 1)

    for id_member in list2:
        if scores[id_member] > 0.5:
            # print(1)
            yes.append(id_member)
            continue
        else:
            for topic_member in members[id_member]:
                if members[id_member][topic_member] > 3 and topic_member in groups[id_group]['topic']:
                    yes.append(id_member)
                    break
            for topic_member in members[id_member]:
                if members[id_member][topic_member] < -3 and topic_member in groups[id_group]['topic']:
                    no.append(id_member)
                    break
            if scores[id_member] > 0.3:
                maybe.append(id_member)
                continue
            no.append(id_member)
    return yes, maybe, no


def predicts(members, groups, events_val, pattern, member_events, events_item):
    folds_num = 5
    sum_y = 0  # 370950
    sum_m = 0  # 5334
    sum_n = 0  # 272125
    w_y = r_y = 0
    w_m = r_m = 0
    w_n = r_n = 0
    total = len(groups)
    count = 0
    for id_group in groups:
        # print("start predict" + id_group)
        i = 0
        # 进度显示
        if count + 1 == total:
            percent = 100.0
            print('predict_Progress: %s [%d/%d]' % (str(percent) + '%', count + 1, total), end='\n')
        else:
            percent = round(1.0 * count / total * 100, 2)
            print('predict_Progress: %s [%d/%d]' % (str(percent) + '%', count + 1, total), end='\r')
        count = count + 1

        for id_event in events_val[id_group]:
            (yes, maybe, no) = predict(id_group, id_event, members, groups, events_val, pattern, member_events, events_item)
            for id_member in events_val[id_group][id_event]['yes']:
                sum_y = sum_y + 1
                if id_member in yes:
                    r_y = r_y + 1
                elif id_member in maybe:
                    w_m = w_m + 1
                else:
                    w_n = w_n + 1
            for id_member in events_val[id_group][id_event]['no']:
                sum_n = sum_n + 1
                if id_member in yes:
                    w_y = w_y + 1
                elif id_member in maybe:
                    w_m = w_m + 1
                else:
                    r_n = r_n + 1
            for id_member in events_val[id_group][id_event]['maybe']:
                sum_m = sum_m + 1
                if id_member in yes:
                    w_y = w_y + 1
                elif id_member in maybe:
                    r_m = r_m + 1
                else:
                    w_n = w_n + 1
            i = (i + 1) % folds_num
    r = [pattern,
         r_y / sum_y, r_y / (r_y + w_y),
         r_m / sum_m, r_m / (r_m + w_m),
         r_n / sum_n, r_n / (r_n + w_n),
         (r_y + r_m + r_n) / (sum_y + sum_m + sum_n)]
    print('yes召回率：', r_y / sum_y)
    print('yes正确率：', r_y / (r_y + w_y))
    print('maybe召回率：', r_m / sum_m)
    print('maybe正确率：', r_m / (r_m + w_m))
    print('no召回率：', r_n / sum_n)
    print('no正确率：', r_n / (r_n + w_n))
    # print('总召回率：', (r_y + r_m + r_n) / (sum_y + sum_m + sum_n), pattern)
    print('总正确率：', (r_y + r_m + r_n) / (sum_y + sum_m + sum_n), pattern)
    return r


def main():
    global net_weights
    members = read_file.get_members()
    groups = read_file.get_groups()
    event_records = read_file.get_events()
    wb = Workbook()
    ws = wb.active
    sr1 = 2
    sl1 = 2
    ws.cell(sr1-1, sl1).value = 'pattern'
    ws.cell(sr1 - 1, sl1+1).value = 'yes召回率'
    ws.cell(sr1 - 1, sl1+2).value = 'yes正确率'
    ws.cell(sr1 - 1, sl1+3).value = 'maybe召回率'
    ws.cell(sr1 - 1, sl1+4).value = 'maybe正确率'
    ws.cell(sr1 - 1, sl1+5).value = 'no召回率'
    ws.cell(sr1 - 1, sl1+6).value = 'no正确率'
    # print('总召回率：', (r_y + r_m + r_n) / (sum_y + sum_m + sum_n), pattern)
    ws.cell(sr1 - 1, sl1+7).value = '总正确率'
    for val_id in range(0, 5):
        events_val, events_train = split.get_folds(val_id, members, event_records, groups)
        members_prefer = t.get_member_habits(1, events_train)
        print("start predict")
        '''
        pattern 0: 只考虑成员初始兴趣
        pattern 1: 考虑训练集全部数据，对成员兴趣进行调整
        pattern 2: 考虑训练集部分数据，成员兴趣的调整只依赖于所预测时间点以前的数据 

        pattern 3: 考虑训练集全部数据，对成员兴趣进行调整，引入网络，权值为相似度（即两个人在该社团的兴趣相似度）
        pattern 4: 考虑训练集部分数据，成员（只对所要预测的进行调整）兴趣的调整只依赖于所预测时间点以前的数据，引入网络，权值为相似度（即两个人在该社团的兴趣相似度）
        pattern 5: 考虑训练集全部数据，对成员兴趣进行调整，引入相似度网络，当相似度超过阈值时建立边
        pattern 6: 考虑训练集全部数据，对成员兴趣进行调整，引入网络，当以往两个人出现在同一回应结果中是有边
        pattern 7: 考虑训练集全部数据，对成员兴趣进行调整，引入网络，边权值为两个人出现的次数比上两个人出现的总次数
        
        '''
        ws.cell(sr1 + (val_id * 10), sl1).value = '结果' + str(val_id)
        for pattern in range(5, 8):
            '''
            if pattern == 2 or pattern == 4:
                continue
            '''
            if pattern != 0 and pattern != 2 and pattern != 4:
                members = split.members_update_by_train(members, groups, events_train)
            member_events = t.get_member_events(events_train)
            # print(member_events)
            events_item = t.get_events_item(events_train, members, groups)
            if pattern == 6:
                net_weights = g.build_graph_1(val_id)
            elif pattern == 7:
                net_weights = g.build_graph_2(val_id)
            r = predicts(members, groups, events_val, pattern, member_events, events_item)
            for l in range(0, 8):
                ws.cell(sr1 + val_id + (10 * pattern), sl1 + l).value = r[l]
        '''
        net_weights = g.build_graph_2(val_id)
        predicts(members, groups, events_val, events_train, pattern=5)
        '''
    wb.save('result567.xlsx')


if __name__ == '__main__':
    main()
    print('Done')

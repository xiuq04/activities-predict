#!/usr/bin/env python
# -*- coding: utf-8 -*-
from math import sqrt
import random
import read_file
import split
import timeitem as t
import build_graph as g

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

net_weights = g.Graph()
mode = 1


def scores_comp(list1, list2, group_topic, members):
    scores = {}
    for id_member in list2:
        same = 0
        num = 0
        if id_member in list1:
            same = len(group_topic)
        for topic_member in group_topic:
            num = num + 1
            if topic_member in members[id_member]:
                same = same + members[id_member][topic_member]
        scores[id_member] = same / num
    return scores


def sims_comp(list1, list2, members, group_topics):
    sims = {}
    add = []
    for id_member in list(set(list1 + list2)):
        if id_member not in members:
            for topic in group_topics:
                members[id_member][topic] = 1
            add.append(id_member)
    avg = {}
    for id_member in list(set(list1 + list2)):
        num = 0
        avg[id_member] = 0
        for topic in group_topics:
            if topic in members[id_member]:
                avg[id_member] += members[id_member][topic]
                num += 1
        '''
        for topic in members[id_member]:
            avg[id_member] += members[id_member][topic]
            num += 1
        '''
        avg[id_member] /= (num if num != 0 else 1)
    for id_member2 in list2:
        sim = {}
        for id_member1 in list1:
            mul = 0
            len1 = 0
            len2 = 0
            for topic in group_topics:
                # for topic in members[id_member2]:
                if topic not in members[id_member1]:
                    tmp1 = avg[id_member1]
                else:
                    tmp1 = members[id_member1][topic]
                if topic not in members[id_member2]:
                    tmp2 = avg[id_member2]
                else:
                    tmp2 = members[id_member2][topic]
                mul += (tmp1 - avg[id_member1]) * (tmp2 - avg[id_member2])
                len1 += (tmp1 - avg[id_member1]) * (tmp1 - avg[id_member1])
                len2 += (tmp2 - avg[id_member2]) * (tmp2 - avg[id_member2])
            tmp = sqrt(len1 * len2)
            sim[id_member1] = mul / (tmp if tmp != 0 else 1)
        sims[id_member2] = sim
    for id_member in add:
        del members[id_member]
    # return sorted(sim.items(), key=lambda x: x[1], reverse=False)
    return sims


def weights_to_sims(members):
    sims = {}
    for id_member1 in members:
        sim = {}
        for id_member2 in members:
            if id_member2 != id_member1:

                if id_member1 != 'null' and id_member2 != 'null' and int(id_member1[1:]) in net_weights.vertList \
                        and int(id_member2[1:]) in net_weights.vertList \
                        and net_weights.vertList[int(id_member2[1:])] \
                        in net_weights.vertList[int(id_member1[1:])].connectedTo.keys():
                    sim[id_member2] = net_weights.similarityCalc(int(id_member1[1:]),
                                                                 net_weights.vertList[int(id_member2[1:])])
                else:
                    sim[id_member2] = 0
        sims[id_member1] = sim
    return sims


def form(value):
    if value > 0.5:
        return 1
    elif value < 0.3:
        return 0
    else:
        return 0.5


def scores_high_comp(scores_base, list2, sims,epoch):
    h = {}
    scores = scores_base
    for r in range(0, epoch):
        for id_member1 in list2:
            h[id_member1] = 0
            sim = sims[id_member1]
            for id_member2 in list2:
                if id_member2 != id_member1:
                    h[id_member1] += sim[id_member2] * (0.5 - form(scores[id_member2]))
            scores[id_member1] = form(scores[id_member1] - h[id_member1])
    return scores


def predict(id_group, id_event, members, groups, events_val, events_train, members_prefer, epoch):
    yes = []
    maybe = []
    no = []
    list1 = events_val[id_group][id_event]['org']
    list2 = events_val[id_group][id_event]['yes'] + events_val[id_group][id_event]['maybe'] + \
            events_val[id_group][id_event]['no']
    scores_base = scores_comp(list1, list2, groups[id_group]['topic'], members)
    sims = weights_to_sims(list2)
    scores = scores_high_comp(scores_base, list2, sims,epoch)
    
    for id_member in list2:
        if id_member not in members_prefer or id_group not in members_prefer[id_member]:
            prefer_yes = 0
            prefer_maybe = 0
            prefer_no = 0
        else:
            total = members_prefer[id_member][id_group][0] + members_prefer[id_member][id_group][1] \
                    + members_prefer[id_member][id_group][2]
            prefer_yes = members_prefer[id_member][id_group][0] / total - 1 / 3
            prefer_maybe = members_prefer[id_member][id_group][1] / total - 1 / 3
            prefer_no = members_prefer[id_member][id_group][2] / total - 1 / 3

        if form(scores[id_member] + prefer_yes) == 1:
            yes.append(id_member)
        elif form(scores[id_member] - prefer_no) == 0:
            no.append(id_member)
        else:
            maybe.append(id_member)
    return yes, maybe, no


def predicts(members, groups, events_val, events_train,members_prefer, epoch):
    folds_num = 5
    sum_y = 0  # 370950
    sum_m = 0  # 5334
    sum_n = 0  # 272125
    w_y = r_y = 0
    w_m = r_m = 0
    w_n = r_n = 0
    total = len(groups)
    count = 0
    for id_group in groups:
        # print("start predict" + id_group)
        i = 0
        # 进度显示
        if count + 1 == total:
            percent = 100.0
            print('predict_Progress: %s [%d/%d]' % (str(percent) + '%', count + 1, total), end='\n')
        else:
            percent = round(1.0 * count / total * 100, 2)
            print('predict_Progress: %s [%d/%d]' % (str(percent) + '%', count + 1, total), end='\r')
        count = count + 1

        for id_event in events_val[id_group]:
            (yes, maybe, no) = predict(id_group, id_event, members, groups, events_val, events_train, members_prefer,epoch)
            for id_member in events_val[id_group][id_event]['yes']:
                sum_y = sum_y + 1
                if id_member in yes:
                    r_y = r_y + 1
                elif id_member in maybe:
                    w_m = w_m + 1
                else:
                    w_n = w_n + 1
            for id_member in events_val[id_group][id_event]['no']:
                sum_n = sum_n + 1
                if id_member in yes:
                    w_y = w_y + 1
                elif id_member in maybe:
                    w_m = w_m + 1
                else:
                    r_n = r_n + 1
            for id_member in events_val[id_group][id_event]['maybe']:
                sum_m = sum_m + 1
                if id_member in yes:
                    w_y = w_y + 1
                elif id_member in maybe:
                    r_m = r_m + 1
                else:
                    w_n = w_n + 1
            i = (i + 1) % folds_num
    r = [epoch,
         r_y / sum_y, r_y / (r_y + w_y),
         r_m / sum_m, r_m / (r_m + w_m),
         r_n / sum_n, r_n / (r_n + w_n),
         (r_y + r_m + r_n) / (sum_y + sum_m + sum_n)]
    print('yes召回率：', r_y / sum_y)
    print('yes正确率：', r_y / (r_y + w_y))
    print('maybe召回率：', r_m / sum_m)
    print('maybe正确率：', r_m / (r_m + w_m))
    print('no召回率：', r_n / sum_n)
    print('no正确率：', r_n / (r_n + w_n))
    # print('总召回率：', (r_y + r_m + r_n) / (sum_y + sum_m + sum_n), pattern)
    print('总正确率：', (r_y + r_m + r_n) / (sum_y + sum_m + sum_n), epoch)
    return r


def main():
    global net_weights
    members = read_file.get_members()
    groups = read_file.get_groups()
    event_records = read_file.get_events()
    wb = Workbook()
    ws = wb.active
    sr1 = 2
    sl1 = 2
    ws.cell(sr1-1, sl1).value = 'epoch'
    ws.cell(sr1 - 1, sl1+1).value = 'yes召回率'
    ws.cell(sr1 - 1, sl1+2).value = 'yes正确率'
    ws.cell(sr1 - 1, sl1+3).value = 'maybe召回率'
    ws.cell(sr1 - 1, sl1+4).value = 'maybe正确率'
    ws.cell(sr1 - 1, sl1+5).value = 'no召回率'
    ws.cell(sr1 - 1, sl1+6).value = 'no正确率'
    # print('总召回率：', (r_y + r_m + r_n) / (sum_y + sum_m + sum_n), pattern)
    ws.cell(sr1 - 1, sl1+7).value = '总正确率'
    for val_id in range(0, 5):
        events_val, events_train = split.get_folds(val_id, members, event_records, groups)
        members_prefer = t.get_member_habits(mode, events_train)
        ws.cell(sr1 + (val_id * 12), sl1).value = '结果' + str(val_id)
        print("start predict")

        net_weights = g.build_graph_1(val_id)
        members = split.members_update_by_train(members, groups, events_train)
        for epoch in range(0, 11):
            r = predicts(members, groups, events_val, events_train, members_prefer, epoch)
            for i in range(0, 8):
                ws.cell(sr1 + (val_id * 12) + epoch, sl1 + i).value = r[i]
    wb.save('result-sim.xlsx')


if __name__ == '__main__':
    main()
    print('Done')
